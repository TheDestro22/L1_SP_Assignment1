# Program: Application1 (Population Application) A program that manages population statistics for countries.

# Input format:
#   User enters data for countries, including:
#   Country name
#   List of state/province/governorate names and their populations (one pair per line).

#   Program Functionalities:
#       Load data: Add new countries with their state/province/governorate populations.
#       View data: See populations of each state/province/governorate and the whole country.
#       Find extremes: Discover the state/province/governorate with the highest and lowest population.
#       Save data: Stores everything in an Excel file for later use.
#       Simple to use: Prompts guide you through data entry and operations.


# Authors: 
#       Gamal Megahed Sayed:    Section: NA    ID:20231039
#       Mohamed Ehab Sabry:     Section: NA    ID:20230547
#       Ayman Ahmed Abdelsamie: Section: NA    ID:20230080

# Version: 1.0
# Date: 1/3/2024






from openpyxl import Workbook

def get_int(message=""):
    # Function to get an integer input from the user
    valid_input = False
    while not valid_input:
        try:
            num = int(input(message))
            valid_input = True
        except:
            print("\033[1;91mInvalid number\033[0m")
    return num

def load(population_data, wb):
    # Function to load population data for a country
    country = input("Enter The Country Name: ")
    sheet = wb.create_sheet(title=country)
    sheet['A1'] = 'city'
    sheet['B1'] = 'population'
    n = get_int("How Many state / province / governorate's you want to add: ")
    while n < 0:
        print("\033[1;91mInvalid number\033[0m")
        n = get_int("How Many state / province / governorate's you want to add: ")

    country_data = {}
    for i in range(n):
        city = input("Enter state / province / governorate name: ")
        population = get_int("Enter population: ")
        while population < 0:
            print("\033[1;91mInvalid number\033[0m")
            population = get_int("Enter population: ")
        country_data[city] = population
        sheet.append([city, population])

    population_data[country] = country_data
    wb.save('population_statistics.xlsx')
    return population_data

def display_population_data(population_data):
    # Function to display population data for a country
    sum = 0
    country = input("Enter The Country Name: ")
    while country not in population_data.keys():
        print("\033[1;91mInvalid Country Name\033[0m")
        country = input("Enter The Country Name: ")

    for key, item in population_data[country].items():
        print(f"{key} : {item}")
        sum += item

    print(f"Total Population: {sum}")
    return population_data

def get_key(value, population_data):
    # Function to get the key for a given value in population data
    for country, cities in population_data.items():
        for city, population in cities.items():
            if population == value:
                return city

def display_population_data_highest_lowest(population_data):
    # Function to display the state/province/governorate with the highest and lowest population in a country
    max_population = 0
    min_population = 37800000
    country = input("Enter The Country Name: ")
    while country not in population_data.keys():
        print("\033[1;91mInvalid Country Name\033[0m")
        country = input("Enter The Country Name: ")

    for key, item in population_data[country].items():
        min_population = min(min_population, item)
        max_population = max(max_population, item)

    key_min = get_key(min_population, population_data)
    key_max = get_key(max_population, population_data)

    print(f"most population country:\n{key_max}:{max_population} ")
    print(f"lowest population country:\n{key_min}:{min_population} ")
    return population_data

# Initial population data
population_data = {
    'USA': {
        'Alabama': 5024279,
        'Alaska': 733391,
        'Arizona': 7151502,
        'California': 39538223,
        'Florida': 21477737,
        'Georgia': 10617423,
        'Illinois': 12671821,
        'New York': 19453561,
        'Texas': 28995881,
        'Washington': 7693612,
        'Ohio': 11747694,
        'Michigan': 10045029,
        'Pennsylvania': 12820878,
        'Virginia': 8535519,
        'Colorado': 5758736,
    },
    'Canada': {
        'Ontario': 14677900,
        'Quebec': 8494500,
        'British Columbia': 5103500,
        'Alberta': 4413146,
        'Manitoba': 1379584,
        'Saskatchewan': 1181666,
        'Nova Scotia': 979115,
        'New Brunswick': 781476,
        'Newfoundland and Labrador': 521365,
        'Prince Edward Island': 159625,
        'Northwest Territories': 44904,
        'Yukon': 42176,
        'Nunavut': 39285,
    },
    'Egypt': {
        'Cairo': 10030000,
        'Giza': 9200000,
        'Sharkia': 7640000,
        'Alexandria': 5163750,
        'Aswan': 1473975,
        'Luxor': 1250208,
        'Port Said': 755994,
        'Suez': 764757,
        'Ismailia': 1725354,
        'Damietta': 1770821,
        'Minya': 5801584,
        'Faiyum': 4310521,
        'Beni Suef': 3176574,
        'Asyut': 4259670,
        'Sohag': 4967400,
    },
    'Germany': {
        'Berlin': 3669491,
        'Hamburg': 1841179,
        'Bavaria': 13124737,
        'North Rhine-Westphalia': 17932651,
        'Lower Saxony': 7982448,
        'Baden-Württemberg': 11100394,
        'Hesse': 6288080,
        'Saxony': 4071971,
        'Thuringia': 2133378,
        'Brandenburg': 2521893,
    },
    'France': {
        'Île-de-France': 12278289,
        'Auvergne-Rhône-Alpes': 8106706,
        'Hauts-de-France': 5971369,
        'Provence-Alpes-Côte d\'Azur': 5100779,
        'Grand Est': 5524192,
        'Pays de la Loire': 3882936,
        'Brittany': 3185692,
        'Nouvelle-Aquitaine': 6079610,
        'Occitanie': 6003941,
        'Centre-Val de Loire': 2572815,
    },
    'Italy': {
        'Lombardy': 10060574,
        'Lazio': 5879082,
        'Campania': 5801692,
        'Sicily': 4999891,
        'Veneto': 4905832,
        'Emilia-Romagna': 4467118,
        'Piedmont': 4356406,
        'Apulia': 4029053,
        'Tuscany': 3729641,
        'Calabria': 1947131,
    },
    'Spain': {
        'Andalusia': 8446561,
        'Catalonia': 7675217,
        'Community of Madrid': 6663394,
        'Valencian Community': 5003769,
        'Castile and León': 2399544,
        'Galicia': 2699499,
        'Basque Country': 2180300,
        'Castilla-La Mancha': 2032863,
        'Canary Islands': 2229010,
        'Aragon': 1319291,
    },
    'Saudi Arabia': {
        'Riyadh': 8114773,
        'Makkah': 7974576,
        'Eastern Province': 4896838,
        'Asir': 2288282,
        'Madinah': 2284304,
        'Qassim': 1446096,
        'Jizan': 1645513,
        'Hail': 706436,
        'Najran': 761779,
        'Tabuk': 970858,
    },
    'United Arab Emirates': {
        'Dubai': 3133326,
        'Abu Dhabi': 2842354,
        'Sharjah': 1446365,
        'Ajman': 513423,
        'Fujairah': 152779,
        'Ras Al Khaimah': 451587,
        'Umm Al Quwain': 84600,
    },
    'Russia': {
        'Moscow': 12615882,
        'St. Petersburg': 5398064,
        'Novosibirsk': 1625631,
        'Yekaterinburg': 1493749,
        'Nizhny Novgorod': 1259013,
        'Kazan': 1257661,
        'Chelyabinsk': 1205473,
        'Omsk': 1172070,
        'Samara': 1156659,
        'Rostov-on-Don': 1125291,
    },
    'Brazil': {
        'São Paulo': 12252023,
        'Rio de Janeiro': 6718903,
        'Brasília': 3055149,
        'Salvador': 2886698,
        'Fortaleza': 2669342,
        'Belo Horizonte': 2512070,
        'Manaus': 2182763,
        'Curitiba': 1933105,
        'Recife': 1645727,
        'Goiânia': 1516113,
    },
    'India': {
        'Mumbai': 12691836,
        'Delhi': 11007835,
        'Bangalore': 8443675,
        'Hyderabad': 6809970,
        'Ahmedabad': 5577940,
        'Chennai': 4681087,
        'Kolkata': 4486679,
        'Surat': 4478297,
        'Pune': 3124456,
        'Jaipur': 3073350,
    },
    'China': {
        'Shanghai': 27178579,
        'Beijing': 21706919,
        'Guangzhou': 15037594,
        'Shenzhen': 13430905,
        'Tianjin': 13214994,
        'Wuhan': 11508141,
        'Dongguan': 8660293,
        'Chengdu': 8312605,
        'Nanjing': 8115050,
        'Hong Kong': 7481500,
    },
    'Japan': {
        'Tokyo': 13973140,
        'Yokohama': 3779758,
        'Osaka': 2802138,
        'Nagoya': 2327731,
        'Sapporo': 1973115,
        'Fukuoka': 1597942,
        'Kawasaki': 1538644,
        'Kobe': 1523195,
        'Kyoto': 1474570,
        'Saitama': 1369424,
    },
    'Australia': {
        'Sydney': 5312163,
        'Melbourne': 5078193,
        'Brisbane': 2482162,
        'Perth': 2083271,
        'Adelaide': 1345774,
        'Gold Coast': 679127,
        'Newcastle': 493697,
        'Canberra': 460916,
        'Wollongong': 302739,
        'Hobart': 240342,
    },
    'South Korea': {
        'Seoul': 9720846,
        'Busan': 3413841,
        'Incheon': 2957026,
        'Daegu': 2428226,
        'Daejeon': 1475226,
        'Gwangju': 1455059,
        'Suwon': 1202123,
        'Ulsan': 1169763,
        'Bucheon': 859406,
        'Seongnam': 975892,
    },
    'Mexico': {
        'Mexico City': 9209944,
        'Ecatepec': 17427790,
        'Guadalajara': 14601408,
        'Puebla': 6840673,
        'Tijuana': 1856705,
        'Monterrey': 1342744,
        'León': 1238962,
        'Zapopan': 1227803,
        'Juárez': 1217818,
        'Tlalnepantla': 1207240,
    },
    'Indonesia': {
        'Jakarta': 10770487,
        'Surabaya': 2965619,
        'Bandung': 2894150,
        'Medan': 2652434,
        'Semarang': 1642648,
        'Makassar': 1541846,
        'Palembang': 1479619,
        'Tangerang': 1434284,
        'Depok': 1309123,
        'Manado': 1199549,
    },
    'Nigeria': {
        'Lagos': 14368410,
        'Kano': 5409010,
        'Ibadan': 3389863,
        'Kaduna': 1654391,
        'Port Harcourt': 1604239,
        'Benin City': 1471480,
        'Maiduguri': 1346419,
        'Zaria': 1218663,
        'Jos': 1153905,
        'Owerri': 1106832,
    },
    'Kenya': {
    'Central Province': 4435801,
    'Coast Province': 4032482,
    'Eastern Province': 5641429,
    'Nairobi Province': 3218070,
    'North Eastern Province': 2520593,
    'Nyanza Province': 5005140,
    'Rift Valley Province': 10006804,
    'Western Province': 4324000,
    'Former Eastern Province': 1308758,
    'Former Central Province': 776215,
  },
  'Iran': {
    'Tehran Province': 15021846,
    'Isfahan Province': 5121200,
    'Khuzestan Province': 4884668,
    'Razavi Khorasan Province': 8021747,
    'Fars Province': 4851555,
    'Mazandaran Province': 3612089,
    'East Azerbaijan Province': 3909457,
    'West Azerbaijan Province': 3080574,
    'Kermanshah Province': 1949788,
    'Sistan and Baluchestan Province': 2775014,
  },
  'Colombia': {
    'Cundinamarca Department': 11017430,  # Bogotá included
    'Antioquia Department': 6607390,
    'Valle del Cauca Department': 4414255,
    'Santander Department': 2088953,
    'Bolívar Department': 2127724,
    'Atlántico Department': 2493877,
    'Tolima Department': 1433010,
    'Nariño Department': 1383327,
    'Caldas Department': 987235,
    'Meta Department': 1036283,
  }
}

# Create a workbook
wb = Workbook()

# Create sheets for each country and populate them with data
for country, city in population_data.items():
    sheet = wb.create_sheet(title=country)
    sheet['A1'] = 'city'
    sheet['B1'] = 'population'
    for c, p in city.items():
        sheet.append([c, p])

# Remove the default sheet created by openpyxl
wb.remove(wb.active)
wb.save('population_statistics.xlsx')

# Main program loop
choose = True
while choose:
    choose = input("What do you want to do?\n(1) Enter a new country to load\n(2) Display the population of each state / province / governorate and total population of the country\n(3) Display the state / province / governorate with the highest population and the one with the lowest population.\n(4) Exit the program\n\nChoice: ")

    while choose not in {'1', '2', '3', '4'}:
        print("\033[1;91mInvalid choice. Please try again\033[0m")
        choose = input("What do you want to do?\n(1) Enter a new country to load\n(2) Display the population of each state / province / governorate and total population of the country\n(3) Display the state / province / governorate with the highest population and the one with the lowest population.\n(4) Exit the program\n\nChoice: ")

    if choose == '1':
        load(population_data, wb)
    elif choose == '2':
        population_data = display_population_data(population_data)
    elif choose == '3':
        population_data = display_population_data_highest_lowest(population_data)
    elif choose == '4':
        print("Thanks for using our program! 🥰")
        break
